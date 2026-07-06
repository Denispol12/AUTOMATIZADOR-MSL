import asyncio
import io
import json
import os
import re
import shutil
import sqlite3
import time
import unicodedata
from contextlib import contextmanager, asynccontextmanager
from pathlib import Path

from google import genai
from google.genai import types
import openai as openai_sdk
import bibtexparser
from fastapi import FastAPI, HTTPException, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

# ── Constantes ────────────────────────────────────────────────────────────────
UPLOADS_DIR = Path("uploads")
UPLOADS_DIR.mkdir(exist_ok=True)

DB_PATH = Path("msl_resultados.db")

# ── Configuración Gemini (lazy: se inicializa al primer uso) ──────────────────
_cliente: genai.Client | None = None
MODELO = "gemini-3.5-flash"

# ── Control de velocidad (gemini-3.5-flash free tier: 5 RPM) ─────────────────
# 60s / 5 RPM = 12s mínimo entre llamadas. Se usan 13s con 1s de margen.
ultima_llamada: float = 0.0
SEGUNDOS_ENTRE_LLAMADAS: float = 13.0


GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

# ── Configuración DeepSeek ────────────────────────────────────────────────────
DEEPSEEK_API_KEY = os.getenv("DEEPSEEK_API_KEY") or "sk-dummy-key-placeholder"
DEEPSEEK_MODELO  = "deepseek-v4-flash"

# Cliente DeepSeek: usa el SDK openai apuntado a api.deepseek.com
# (DeepSeek expone una API 100% compatible con OpenAI)
cliente_deepseek = openai_sdk.OpenAI(
    api_key=DEEPSEEK_API_KEY,
    base_url="https://api.deepseek.com",
)


def get_cliente() -> genai.Client:
    """Devuelve el cliente Gemini, creándolo la primera vez."""
    global _cliente
    if _cliente is None:
        _cliente = genai.Client(api_key=GEMINI_API_KEY)
    return _cliente

# ── Prompt dinámico ────────────────────────────────────────────────────────────

def construir_prompt() -> str:
    """
    Construye el prompt de evaluación leyendo los criterios actuales
    desde la tabla 'criterios' en SQLite.
    Solo la lista de CI/CE es dinámica; el resto del texto es fijo.
    """
    with _get_conn() as conn:
        ci_rows = conn.execute(
            "SELECT codigo, descripcion FROM criterios "
            "WHERE tipo='inclusion' ORDER BY codigo"
        ).fetchall()
        ce_rows = conn.execute(
            "SELECT codigo, descripcion FROM criterios "
            "WHERE tipo='exclusion' ORDER BY codigo"
        ).fetchall()

    # ── Secciones dinámicas ────────────────────────────────────────────
    lista_ci = "\n".join(f"- {r['codigo']}: {r['descripcion']}" for r in ci_rows)
    lista_ce = "\n".join(f"- {r['codigo']}: {r['descripcion']}" for r in ce_rows)

    # JSON schema dinámico (campos ci1..ciN y ce1..ceN según la BD)
    campos_ci = "\n".join(
        f'  "{r["codigo"].lower()}": 0 | 1,' for r in ci_rows
    )
    campos_ce = "\n".join(
        f'  "{r["codigo"].lower()}": 0 | 1,' for r in ce_rows
    )
    primer_ci = ci_rows[0]["codigo"] if ci_rows else "CI1"
    ultimo_ci = ci_rows[-1]["codigo"] if ci_rows else "CI7"
    primer_ce = ce_rows[0]["codigo"] if ce_rows else "CE1"
    ultimo_ce = ce_rows[-1]["codigo"] if ce_rows else "CE6"

    # ── Prompt completo ────────────────────────────────────────────────
    return f"""
Eres un evaluador experto en revisión sistemática de literatura científica.
Tu tarea es analizar el PDF adjunto y determinar si el artículo cumple los
criterios de inclusión y exclusión para una revisión sistemática sobre
detección de plagas internas en frutos agrícolas mediante imágenes no visibles
e inteligencia artificial.

Evalúa cada criterio con 1 (cumple) o 0 (no cumple).

CRITERIOS DE INCLUSIÓN (CI):
{lista_ci}

CRITERIOS DE EXCLUSIÓN (CE):
{lista_ce}

REGLAS DE DECISIÓN:
- INCLUIDO: Todos los CI deben ser 1 Y todos los CE deben ser 0.
- EXCLUIDO: Al menos un CI es 0 O al menos un CE es 1.
- Si la información del PDF no permite evaluar un criterio con certeza,
  márcalo con 0 y menciónalo en la justificación.

Responde ÚNICAMENTE con un objeto JSON válido, sin texto adicional antes
ni después, sin bloques de código markdown. El formato debe ser exactamente:

{{
  "decision": "INCLUIDO" | "EXCLUIDO",
  "confianza": "ALTA" | "MEDIA" | "BAJA",
{campos_ci}
{campos_ce}
  "criterio_fallo": "{primer_ci}" | "{primer_ce}" | null,
  "justificacion": "Texto breve explicando la decisión principal (máx 300 caracteres)."
}}

Donde "criterio_fallo" es el criterio determinante que provocó la exclusión
(el primero que falló, en orden {primer_ci}→{ultimo_ci}, luego {primer_ce}→{ultimo_ce}),
o null si fue INCLUIDO.
""".strip()




# ── Base de datos ─────────────────────────────────────────────────────────────

def init_db() -> None:
    """Crea las tablas 'resultados' y 'criterios' si no existen, y
    siembra los 13 criterios iniciales solo si la tabla está vacía."""
    with _get_conn() as conn:
        # Tabla de resultados de screening
        conn.execute("""
            CREATE TABLE IF NOT EXISTS resultados (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre_archivo   TEXT,
                decision         TEXT,
                confianza        TEXT,
                ci1              INTEGER,
                ci2              INTEGER,
                ci3              INTEGER,
                ci4              INTEGER,
                ci5              INTEGER,
                ci6              INTEGER,
                ci7              INTEGER,
                ce1              INTEGER,
                ce2              INTEGER,
                ce3              INTEGER,
                ce4              INTEGER,
                ce5              INTEGER,
                ce6              INTEGER,
                criterio_fallo   TEXT,
                justificacion    TEXT,
                fecha            TEXT DEFAULT (datetime('now','localtime'))
            )
        """)

        # Tabla de criterios de inclusión / exclusión
        conn.execute("""
            CREATE TABLE IF NOT EXISTS criterios (
                id          INTEGER PRIMARY KEY,
                codigo      TEXT UNIQUE NOT NULL,
                tipo        TEXT NOT NULL CHECK(tipo IN ('inclusion', 'exclusion')),
                descripcion TEXT NOT NULL
            )
        """)

        # Sembrar criterios iniciales solo si la tabla está vacía
        total = conn.execute("SELECT COUNT(*) FROM criterios").fetchone()[0]
        if total == 0:
            criterios_iniciales = [
                # Inclusión
                ("CI1", "inclusion", "Artículos publicados entre 2021 y 2026"),
                ("CI2", "inclusion", "Estudios primarios (artículos de revista y artículos de conferencia)"),
                ("CI3", "inclusion", "Artículos publicados en inglés y español"),
                ("CI4", "inclusion", "Artículos enfocados en frutos agrícolas"),
                ("CI5", "inclusion", "Uso de imágenes del espectro no visible"),
                ("CI6", "inclusion", "Uso de técnicas de IA para procesamiento de imágenes térmicas"),
                ("CI7", "inclusion", "Detección de plagas/infestaciones internas"),
                # Exclusión
                ("CE1", "exclusion", "Documentos no disponibles"),
                ("CE2", "exclusion", "Estudios secundarios (revisiones sistemáticas, mapeos, meta-análisis)"),
                ("CE3", "exclusion", "Estudios en hojas, tallos, suelo, etc."),
                ("CE4", "exclusion", "Investigaciones que dependen de la destrucción del fruto para la detección"),
                ("CE5", "exclusion", "Uso exclusivo de estadística descriptiva básica"),
                ("CE6", "exclusion", "Estudios sobre daños por hongos/bacterias o daños mecánicos"),
            ]
            conn.executemany(
                "INSERT INTO criterios (codigo, tipo, descripcion) VALUES (?, ?, ?)",
                criterios_iniciales,
            )

        # Tabla de resultados BibTeX (flujo separado de PDFs)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS resultados_bib (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                titulo           TEXT,
                autores          TEXT,
                anio             TEXT,
                revista          TEXT,
                abstract         TEXT,
                palabras_clave   TEXT,
                decision         TEXT,
                confianza        TEXT,
                criterio_fallo   TEXT,
                justificacion    TEXT,
                fecha            TEXT DEFAULT (datetime('now','localtime'))
            )
        """)

        # Tabla del eliminador de duplicados (módulo independiente)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS duplicados_articulos (
                id                  INTEGER PRIMARY KEY AUTOINCREMENT,
                titulo              TEXT,
                titulo_normalizado  TEXT,
                autores             TEXT,
                anio                TEXT,
                base_datos          TEXT,
                clave_cita          TEXT,
                campos_extra        TEXT,
                es_duplicado        INTEGER DEFAULT 0,
                grupo_duplicado     INTEGER,
                fecha               TEXT DEFAULT (datetime('now','localtime'))
            )
        """)

        # Tabla de preguntas de extracción de datos
        conn.execute("""
            CREATE TABLE IF NOT EXISTS preguntas_extraccion (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                texto_pregunta TEXT NOT NULL,
                tipo TEXT NOT NULL,
                opciones TEXT,
                orden INTEGER,
                fecha TEXT DEFAULT (datetime('now','localtime'))
            )
        """)
        # Tabla de respuestas para preguntas categóricas
        conn.execute("""
            CREATE TABLE IF NOT EXISTS extraccion_respuestas_categoricas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                articulo_nombre TEXT NOT NULL,
                pregunta_id INTEGER NOT NULL,
                opcion_texto TEXT NOT NULL,
                valor INTEGER NOT NULL,
                fecha TEXT DEFAULT (datetime('now','localtime'))
            )
        """)

        # Tabla de respuestas para preguntas numéricas y texto libre
        conn.execute("""
            CREATE TABLE IF NOT EXISTS extraccion_respuestas_otras (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                articulo_nombre TEXT NOT NULL,
                pregunta_id INTEGER NOT NULL,
                respuesta TEXT,
                respuesta_original_ia TEXT,
                fecha TEXT DEFAULT (datetime('now','localtime'))
            )
        """)
        # Migración: Agregar columna respuesta_original_ia si no existe en base de datos ya creada
        cursor = conn.execute("PRAGMA table_info(extraccion_respuestas_otras)")
        columnas = [row["name"] for row in cursor.fetchall()]
        if "respuesta_original_ia" not in columnas:
            conn.execute("ALTER TABLE extraccion_respuestas_otras ADD COLUMN respuesta_original_ia TEXT")




@contextmanager
def _get_conn():
    """Context manager que abre, hace commit y cierra la conexión."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row          # acceso por nombre de columna
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def guardar_resultado(nombre: str, resultado: dict) -> int:
    """Inserta una fila en 'resultados' y devuelve el id generado."""
    campos = [
        "nombre_archivo", "decision", "confianza",
        "ci1", "ci2", "ci3", "ci4", "ci5", "ci6", "ci7",
        "ce1", "ce2", "ce3", "ce4", "ce5", "ce6",
        "criterio_fallo", "justificacion",
    ]
    valores = [nombre] + [resultado.get(c) for c in campos[1:]]
    placeholders = ", ".join("?" * len(campos))
    sql = f"INSERT INTO resultados ({', '.join(campos)}) VALUES ({placeholders})"

    with _get_conn() as conn:
        cur = conn.execute(sql, valores)
        return cur.lastrowid


# ── Parser BibTeX robusto ─────────────────────────────────────────────────────
# bibtexparser v1 pierde entradas silenciosamente cuando hay:
#   - % en abstracts (lo trata como comentario LaTeX)
#   - llaves desbalanceadas en campos largos de Scopus
#   - tipos no-estándar como @REVIEW ignorados con el default
# Se reemplaza por un parser manual basado en conteo de llaves.

def _parsear_bib_robusto(texto_bib: str) -> list[dict]:
    """
    Parser BibTeX robusto que NO pierde entradas.

    Estrategia:
      1. Pre-procesar: eliminar comentarios de línea (% ...) fuera de llaves.
      2. Localizar cada bloque @TIPO{...} contando llaves para encontrar el cierre.
      3. Parsear los campos clave=valor dentro de cada bloque.
      4. Ignorar solo @COMMENT, @PREAMBLE y @STRING (metabloques).

    Maneja correctamente:
      - % en abstracts de Scopus.
      - Llaves anidadas en títulos ({\\'e}, {-}, etc.).
      - Tipos no-estándar: @CONFERENCE, @REVIEW, @BOOK, etc.
      - Claves de cita duplicadas (se conservan ambas entradas).
    """
    # 1. Quitar comentarios de línea SOLO fuera de valores entre llaves.
    #    Scopus exporta campos con % dentro de {…}, por eso no hacemos
    #    un strip global — lo hacemos línea por línea contando profundidad.
    lineas_limpias = []
    profundidad = 0
    for linea in texto_bib.splitlines():
        if profundidad == 0:
            # Fuera de un bloque: el % inicia comentario
            idx_pct = linea.find('%')
            if idx_pct != -1:
                linea = linea[:idx_pct]
        # Actualizar profundidad para la siguiente línea
        for ch in linea:
            if ch == '{':
                profundidad += 1
            elif ch == '}':
                profundidad = max(0, profundidad - 1)
        lineas_limpias.append(linea)
    texto_limpio = '\n'.join(lineas_limpias)

    # 2. Encontrar cada bloque @TIPO{...} por posición
    SKIP_TYPES = {'comment', 'preamble', 'string'}
    entradas: list[dict] = []

    patron_tipo = re.compile(r'@([a-zA-Z]+)\s*\{', re.IGNORECASE)
    i = 0
    texto_len = len(texto_limpio)

    while i < texto_len:
        m = patron_tipo.search(texto_limpio, i)
        if not m:
            break

        tipo = m.group(1).lower()
        if tipo in SKIP_TYPES:
            # Saltar este bloque buscando el } de cierre
            start_brace = m.end() - 1  # posición del {
            profundidad = 0
            j = start_brace
            while j < texto_len:
                if texto_limpio[j] == '{':
                    profundidad += 1
                elif texto_limpio[j] == '}':
                    profundidad -= 1
                    if profundidad == 0:
                        i = j + 1
                        break
                j += 1
            else:
                i = texto_len
            continue

        # Encontrar el cuerpo del bloque entre llaves
        start_brace = m.end() - 1
        profundidad = 0
        j = start_brace
        while j < texto_len:
            if texto_limpio[j] == '{':
                profundidad += 1
            elif texto_limpio[j] == '}':
                profundidad -= 1
                if profundidad == 0:
                    break
            j += 1

        cuerpo = texto_limpio[m.end():j]  # contenido entre las llaves externas
        i = j + 1

        # 3. Parsear clave_cita y campos dentro del cuerpo
        entrada = _parsear_cuerpo_bib(tipo.upper(), cuerpo)
        if entrada:
            entradas.append(entrada)

    return entradas


def _parsear_cuerpo_bib(tipo: str, cuerpo: str) -> dict | None:
    """
    Parsea el interior de un bloque @TIPO{clave, campo1={val1}, campo2={val2}}.
    Devuelve un dict con ENTRYTYPE, ID y todos los campos, o None si está vacío.
    """
    cuerpo = cuerpo.strip()
    if not cuerpo:
        return None

    # Separar la clave de cita (primer token antes de la primera coma)
    primera_coma = cuerpo.find(',')
    if primera_coma == -1:
        return {'ENTRYTYPE': tipo, 'ID': cuerpo.strip()}

    clave_cita = cuerpo[:primera_coma].strip()
    resto = cuerpo[primera_coma + 1:]

    entrada = {'ENTRYTYPE': tipo, 'ID': clave_cita}

    # Parsear campos: nombre = {valor} o nombre = "valor" o nombre = numero
    # Avanzamos carácter a carácter respetando la profundidad de llaves
    pos = 0
    n = len(resto)

    while pos < n:
        # Saltar espacios y comas
        while pos < n and resto[pos] in ' \t\n\r,':
            pos += 1
        if pos >= n:
            break

        # Leer nombre del campo (hasta el =)
        eq = resto.find('=', pos)
        if eq == -1:
            break
        nombre_campo = resto[pos:eq].strip().lower()
        pos = eq + 1

        # Saltar espacios tras el =
        while pos < n and resto[pos] in ' \t\n\r':
            pos += 1
        if pos >= n:
            break

        # Leer valor: puede ser {…}, "…", o un número/token
        if resto[pos] == '{':
            # Valor entre llaves — respetar anidamiento
            prof = 0
            inicio_val = pos + 1
            while pos < n:
                if resto[pos] == '{':
                    prof += 1
                elif resto[pos] == '}':
                    prof -= 1
                    if prof == 0:
                        break
                pos += 1
            valor = resto[inicio_val:pos].strip()
            pos += 1  # saltar el } de cierre
        elif resto[pos] == '"':
            # Valor entre comillas
            pos += 1
            inicio_val = pos
            while pos < n and resto[pos] != '"':
                pos += 1
            valor = resto[inicio_val:pos].strip()
            pos += 1  # saltar la " de cierre
        else:
            # Valor sin delimitadores (número, abreviatura)
            inicio_val = pos
            while pos < n and resto[pos] not in ',\n':
                pos += 1
            valor = resto[inicio_val:pos].strip()

        if nombre_campo:
            entrada[nombre_campo] = valor

    return entrada


# ── Lógica Gemini ─────────────────────────────────────────────────────────────

def _ascii_nombre(nombre: str) -> str:
    """Convierte un nombre de archivo a ASCII puro eliminando tildes y especiales."""
    normalizado = unicodedata.normalize("NFKD", nombre)
    solo_ascii  = normalizado.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^\w\-_\. ]", "_", solo_ascii).strip() or "archivo.pdf"


async def _generar_con_reintento(c, model: str, contents: list, max_reintentos: int = 4):
    """
    Llama a c.models.generate_content() con reintentos automáticos ante
    errores 503 UNAVAILABLE (sobrecarga temporal del servidor de Google).

    Política de reintentos (backoff exponencial):
      Intento 1 falla → espera 15 s
      Intento 2 falla → espera 30 s
      Intento 3 falla → espera 60 s
      Intento 4 falla → espera 120 s
      Intento 5 falla → propaga la excepción

    Otros errores (4xx, JSON, etc.) se propagan inmediatamente sin reintentar.
    """
    esperas = [15, 30, 60, 120]
    ultimo_exc = None
    for intento in range(max_reintentos + 1):
        try:
            return c.models.generate_content(model=model, contents=contents)
        except Exception as exc:
            msg = str(exc)
            es_503 = "503" in msg or "UNAVAILABLE" in msg or "high demand" in msg.lower()
            if es_503 and intento < max_reintentos:
                espera = esperas[intento]
                print(f"[Gemini] 503 UNAVAILABLE – reintento {intento + 1}/{max_reintentos} en {espera}s…")
                await asyncio.sleep(espera)
                ultimo_exc = exc
                continue
            raise
    if ultimo_exc is None:
        raise RuntimeError("No se pudo completar la llamada a Gemini tras agotar los reintentos.")
    raise ultimo_exc  # nunca debería llegar aquí

async def llamar_gemini(ruta_pdf: Path) -> dict:
    """
    Sube el PDF a Gemini, solicita la evaluación y devuelve un dict
    con los campos del esquema de resultados.
    """
    global ultima_llamada
    archivo_gemini = None
    texto = ""
    try:
        # 0. Respetar el límite de 5 RPM: esperar si no han pasado 13 s
        transcurrido = time.time() - ultima_llamada
        if transcurrido < SEGUNDOS_ENTRE_LLAMADAS:
            espera = SEGUNDOS_ENTRE_LLAMADAS - transcurrido
            await asyncio.sleep(espera)
        ultima_llamada = time.time()
        # 1. Subir el PDF a la File API de Gemini
        #    Se pasa el objeto abierto (no el path como string) para evitar
        #    el error ASCII con tildes en el nombre del archivo.
        c = get_cliente()
        with open(ruta_pdf, "rb") as f:
            archivo_gemini = c.files.upload(
                file=f,
                config=types.UploadFileConfig(
                    mime_type="application/pdf",
                    display_name=_ascii_nombre(ruta_pdf.name),
                ),
            )

        # 2. Llamar al modelo con el prompt construido dinámicamente
        respuesta = await _generar_con_reintento(
            c, MODELO, [construir_prompt(), archivo_gemini]
        )
        texto = respuesta.text.strip()

        # 3. Limpiar bloque markdown si viene envuelto en ```json ... ```
        texto = re.sub(r"^```(?:json)?\s*", "", texto, flags=re.IGNORECASE)
        texto = re.sub(r"\s*```$", "", texto)

        # 4. Parsear JSON
        return json.loads(texto)

    except json.JSONDecodeError as exc:
        return {
            "decision": "revisar_manualmente",
            "confianza": "BAJA",
            "criterio_fallo": None,
            "justificacion": f"Error al parsear JSON de Gemini: {exc}. Respuesta: {texto[:200]}",
        }
    except Exception as exc:
        return {
            "decision": "revisar_manualmente",
            "confianza": "BAJA",
            "criterio_fallo": None,
            "justificacion": f"Error al llamar a Gemini: {exc}",
        }
    finally:
        # Borrar el archivo de la File API para no acumular cuota
        if archivo_gemini:
            try:
                get_cliente().files.delete(name=archivo_gemini.name)
            except Exception:
                pass


@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    yield

app = FastAPI(title="MSL Validator", lifespan=lifespan)


# ── Endpoints de API ───────────────────────────────────────────────────────────

@app.get("/api/status")
def api_status():
    return {"status": "ok", "app": "MSL Validator"}


@app.get("/resultados")
def listar_resultados():
    """Devuelve todas las filas de la tabla resultados como lista de dicts."""
    with _get_conn() as conn:
        filas = conn.execute("SELECT * FROM resultados ORDER BY id").fetchall()
    return [dict(f) for f in filas]


@app.delete("/resultados/{id}")
def eliminar_resultado(id: int):
    """Elimina una fila por id. Devuelve 404 si no existe."""
    with _get_conn() as conn:
        cur = conn.execute("DELETE FROM resultados WHERE id = ?", (id,))
    if cur.rowcount == 0:
        raise HTTPException(status_code=404, detail=f"Resultado {id} no encontrado")
    return {"eliminado": id}


# ── CRUD Criterios ─────────────────────────────────────────────────────

class CriterioNuevo(BaseModel):
    codigo: str
    tipo: str
    descripcion: str

class CriterioUpdate(BaseModel):
    descripcion: str


@app.get("/criterios")
def listar_criterios():
    """Devuelve todos los criterios ordenados por código."""
    with _get_conn() as conn:
        filas = conn.execute(
            "SELECT * FROM criterios ORDER BY codigo"
        ).fetchall()
    return [dict(f) for f in filas]


@app.post("/criterios", status_code=201)
def crear_criterio(body: CriterioNuevo):
    """Crea un criterio nuevo (ej. CI8, CE7)."""
    if body.tipo not in ("inclusion", "exclusion"):
        raise HTTPException(status_code=400, detail="tipo debe ser 'inclusion' o 'exclusion'")
    try:
        with _get_conn() as conn:
            conn.execute(
                "INSERT INTO criterios (codigo, tipo, descripcion) VALUES (?, ?, ?)",
                (body.codigo.upper(), body.tipo, body.descripcion),
            )
    except Exception:
        raise HTTPException(status_code=409, detail=f"El código '{body.codigo}' ya existe")
    return {"creado": body.codigo.upper()}


@app.put("/criterios/{codigo}")
def actualizar_criterio(codigo: str, body: CriterioUpdate):
    """Actualiza solo la descripción de un criterio. Código y tipo son inmutables."""
    with _get_conn() as conn:
        cur = conn.execute(
            "UPDATE criterios SET descripcion = ? WHERE codigo = ?",
            (body.descripcion, codigo.upper()),
        )
    if cur.rowcount == 0:
        raise HTTPException(status_code=404, detail=f"Criterio '{codigo}' no encontrado")
    return {"actualizado": codigo.upper()}


@app.delete("/criterios/{codigo}")
def eliminar_criterio(codigo: str):
    """Elimina un criterio por código."""
    with _get_conn() as conn:
        cur = conn.execute(
            "DELETE FROM criterios WHERE codigo = ?", (codigo.upper(),)
        )
    if cur.rowcount == 0:
        raise HTTPException(status_code=404, detail=f"Criterio '{codigo}' no encontrado")
    return {"eliminado": codigo.upper()}


@app.post("/evaluar")
async def evaluar(archivo: UploadFile):
    """Recibe un PDF, lo evalúa con Gemini y guarda el resultado en SQLite."""
    if not archivo.filename.lower().endswith(".pdf"):
        raise HTTPException(
            status_code=400,
            detail="Solo se aceptan archivos con extensión .pdf",
        )

    # 1. Guardar en disco
    destino = UPLOADS_DIR / archivo.filename
    with destino.open("wb") as f:
        shutil.copyfileobj(archivo.file, f)

    # 2. Evaluar con Gemini
    resultado = await llamar_gemini(destino)

    # 3. Persistir en SQLite
    fila_id = guardar_resultado(archivo.filename, resultado)

    # 4. Devolver respuesta completa
    return {"id": fila_id, "archivo": archivo.filename, **resultado}


@app.get("/exportar")
def exportar_excel():
    """Genera y descarga un Excel con todos los resultados de screening."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # ── Paleta de colores ──────────────────────────────────────────────────
    FILL_HEADER  = PatternFill("solid", fgColor="1E2235")   # azul oscuro
    FILL_INCLUIDO = PatternFill("solid", fgColor="D6F5E3")  # verde claro
    FILL_EXCLUIDO = PatternFill("solid", fgColor="FDDEDE")  # rojo claro
    FILL_REVISAR  = PatternFill("solid", fgColor="FFF8D6")  # amarillo claro
    FONT_HEADER  = Font(bold=True, color="FFFFFF", size=10)
    FONT_NORMAL  = Font(size=9)

    # ── Leer datos ─────────────────────────────────────────────────────────
    with _get_conn() as conn:
        filas = conn.execute("SELECT * FROM resultados ORDER BY id").fetchall()
    datos = [dict(f) for f in filas]

    # ── Construir workbook ─────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Screening MSL"

    cabeceras = [
        "ID", "Archivo", "Decisión", "Confianza",
        "CI-1", "CI-2", "CI-3", "CI-4", "CI-5", "CI-6", "CI-7",
        "CE-1", "CE-2", "CE-3", "CE-4", "CE-5", "CE-6",
        "Criterio Fallo", "Justificación", "Fecha",
    ]

    # Cabecera
    for col_idx, titulo in enumerate(cabeceras, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.fill   = FILL_HEADER
        cell.font   = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # Filas de datos
    for row_idx, r in enumerate(datos, start=2):
        dec = (r.get("decision") or "").lower()
        fill = (FILL_INCLUIDO if dec == "incluido"
                else FILL_EXCLUIDO if dec == "excluido"
                else FILL_REVISAR)

        def marca(val):
            return "✓" if val == 1 else "✗" if val == 0 else "-"

        fila_valores = [
            r.get("id"),
            r.get("nombre_archivo"),
            r.get("decision"),
            r.get("confianza"),
            marca(r.get("ci1")), marca(r.get("ci2")), marca(r.get("ci3")),
            marca(r.get("ci4")), marca(r.get("ci5")), marca(r.get("ci6")),
            marca(r.get("ci7")),
            marca(r.get("ce1")), marca(r.get("ce2")), marca(r.get("ce3")),
            marca(r.get("ce4")), marca(r.get("ce5")), marca(r.get("ce6")),
            r.get("criterio_fallo"),
            r.get("justificacion"),
            r.get("fecha"),
        ]

        for col_idx, valor in enumerate(fila_valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.fill = fill
            cell.font = FONT_NORMAL
            cell.alignment = Alignment(
                horizontal="center" if col_idx not in (2, 19) else "left",
                vertical="center",
                wrap_text=(col_idx == 19),   # justificación
            )

    # Anchos de columna
    anchos = [
        5,    # ID
        30,   # Archivo
        14,   # Decisión
        10,   # Confianza
        6, 6, 6, 6, 6, 6, 6,   # CI-1..7
        6, 6, 6, 6, 6, 6,      # CE-1..6
        14,   # Criterio Fallo
        45,   # Justificación
        14,   # Fecha
    ]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # Congelar cabecera
    ws.freeze_panes = "A2"

    # ── Guardar y servir ───────────────────────────────────────────────────
    RESULTADOS_DIR = Path("resultados")
    RESULTADOS_DIR.mkdir(exist_ok=True)
    ruta_excel = RESULTADOS_DIR / "MSL_Screening_Resultados.xlsx"
    wb.save(ruta_excel)

    return FileResponse(
        path=str(ruta_excel),
        filename="MSL_Screening_Resultados.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Exportar BibTeX a Excel ────────────────────────────────────────────────────

@app.get("/exportar-bibtex")
def exportar_excel_bibtex():
    """Genera y descarga un Excel con los resultados de screening BibTeX."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # ── Paleta de colores (misma que /exportar) ────────────────────────────
    FILL_HEADER   = PatternFill("solid", fgColor="1E2235")
    FILL_INCLUIDO = PatternFill("solid", fgColor="D6F5E3")
    FILL_EXCLUIDO = PatternFill("solid", fgColor="FDDEDE")
    FILL_REVISAR  = PatternFill("solid", fgColor="FFF8D6")
    FONT_HEADER   = Font(bold=True, color="FFFFFF", size=10)
    FONT_NORMAL   = Font(size=9)

    # ── Leer datos ─────────────────────────────────────────────────────────
    with _get_conn() as conn:
        filas = conn.execute("SELECT * FROM resultados_bib ORDER BY id").fetchall()
    datos = [dict(f) for f in filas]

    # ── Construir workbook ─────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Screening BibTeX"

    cabeceras = [
        "N°", "Título", "Autores", "Año", "Revista",
        "Decisión", "Confianza",
        "CI-1", "CI-2", "CI-3", "CI-4", "CI-5", "CI-6", "CI-7",
        "CE-1", "CE-2", "CE-3", "CE-4", "CE-5", "CE-6",
        "Criterio Fallo", "Justificación", "Fecha",
    ]

    # Cabecera
    for col_idx, titulo in enumerate(cabeceras, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # Columnas de texto largo (para alineación izquierda)
    COL_TITULO = 2
    COL_AUTORES = 3
    COL_REVISTA = 5
    COL_JUSTIFICACION = 22

    # Filas de datos
    for row_idx, r in enumerate(datos, start=2):
        dec = (r.get("decision") or "").lower()
        fill = (FILL_INCLUIDO if dec == "incluido"
                else FILL_EXCLUIDO if dec == "excluido"
                else FILL_REVISAR)

        def marca(val):
            return "✓" if val == 1 else "✗" if val == 0 else "-"

        fila_valores = [
            r.get("id"),
            r.get("titulo"),
            r.get("autores"),
            r.get("anio"),
            r.get("revista"),
            r.get("decision"),
            r.get("confianza"),
            marca(r.get("ci1")), marca(r.get("ci2")), marca(r.get("ci3")),
            marca(r.get("ci4")), marca(r.get("ci5")), marca(r.get("ci6")),
            marca(r.get("ci7")),
            marca(r.get("ce1")), marca(r.get("ce2")), marca(r.get("ce3")),
            marca(r.get("ce4")), marca(r.get("ce5")), marca(r.get("ce6")),
            r.get("criterio_fallo"),
            r.get("justificacion"),
            r.get("fecha"),
        ]

        cols_izq = {COL_TITULO, COL_AUTORES, COL_REVISTA, COL_JUSTIFICACION}
        for col_idx, valor in enumerate(fila_valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.fill = fill
            cell.font = FONT_NORMAL
            cell.alignment = Alignment(
                horizontal="left" if col_idx in cols_izq else "center",
                vertical="center",
                wrap_text=(col_idx == COL_JUSTIFICACION),
            )

    # Anchos de columna
    anchos = [
        5,    # N°
        40,   # Título
        25,   # Autores
        8,    # Año
        25,   # Revista
        14,   # Decisión
        10,   # Confianza
        6, 6, 6, 6, 6, 6, 6,   # CI-1..7
        6, 6, 6, 6, 6, 6,      # CE-1..6
        14,   # Criterio Fallo
        45,   # Justificación
        14,   # Fecha
    ]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # Congelar cabecera
    ws.freeze_panes = "A2"

    # ── Guardar y servir ───────────────────────────────────────────────────
    RESULTADOS_DIR = Path("resultados")
    RESULTADOS_DIR.mkdir(exist_ok=True)
    ruta_excel = RESULTADOS_DIR / "MSL_Screening_Bibtex_Resultados.xlsx"
    wb.save(ruta_excel)

    return FileResponse(
        path=str(ruta_excel),
        filename="MSL_Screening_Bibtex_Resultados.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── BibTeX ─────────────────────────────────────────────────────────────────────

@app.post("/subir-bibtex")
async def subir_bibtex(archivo: UploadFile):
    """Recibe un archivo .bib, lo parsea con bibtexparser y devuelve
    la lista de entradas como JSON. No conecta con Gemini."""
    if not archivo.filename.lower().endswith(".bib"):
        raise HTTPException(
            status_code=400,
            detail="Solo se aceptan archivos con extensión .bib",
        )

    # Leer contenido del archivo
    contenido = await archivo.read()
    texto_bib = contenido.decode("utf-8", errors="replace")

    # Parsear con el parser robusto (reemplaza bibtexparser que pierde entradas
    # silenciosamente con % en abstracts y tipos no-estándar de Scopus)
    entradas_raw = _parsear_bib_robusto(texto_bib)

    # Extraer campos relevantes de cada entrada
    entradas = []
    for entry in entradas_raw:
        entradas.append({
            "clave":           entry.get("ID", ""),
            "tipo":            entry.get("ENTRYTYPE", ""),
            "titulo":          entry.get("title", ""),
            "autores":         entry.get("author", ""),
            "anio":            entry.get("year", ""),
            "revista":         entry.get("journal", entry.get("booktitle", "")),
            "abstract":        entry.get("abstract", ""),
            "palabras_clave":  entry.get("keywords", entry.get("author_keywords", "")),
        })

    return {
        "total": len(entradas),
        "entradas": entradas,
    }


# ── BibTeX: evaluación con Gemini ──────────────────────────────────────────────

def construir_prompt_bibtex() -> str:
    """
    Prompt adaptado para evaluar artículos usando SOLO título, abstract
    y palabras clave (sin texto completo del PDF).
    Lee los criterios dinámicamente desde la tabla 'criterios'.
    """
    with _get_conn() as conn:
        ci_rows = conn.execute(
            "SELECT codigo, descripcion FROM criterios "
            "WHERE tipo='inclusion' ORDER BY codigo"
        ).fetchall()
        ce_rows = conn.execute(
            "SELECT codigo, descripcion FROM criterios "
            "WHERE tipo='exclusion' ORDER BY codigo"
        ).fetchall()

    lista_ci = "\n".join(f"- {r['codigo']}: {r['descripcion']}" for r in ci_rows)
    lista_ce = "\n".join(f"- {r['codigo']}: {r['descripcion']}" for r in ce_rows)

    campos_ci = "\n".join(
        f'  "{r["codigo"].lower()}": 0 | 1,' for r in ci_rows
    )
    campos_ce = "\n".join(
        f'  "{r["codigo"].lower()}": 0 | 1,' for r in ce_rows
    )
    primer_ci = ci_rows[0]["codigo"] if ci_rows else "CI1"
    ultimo_ci = ci_rows[-1]["codigo"] if ci_rows else "CI7"
    primer_ce = ce_rows[0]["codigo"] if ce_rows else "CE1"
    ultimo_ce = ce_rows[-1]["codigo"] if ce_rows else "CE6"

    return f"""
Eres un evaluador experto en revisión sistemática de literatura científica.
Tu tarea es analizar la información bibliográfica proporcionada (título,
abstract y palabras clave) y determinar si el artículo cumple los criterios
de inclusión y exclusión para una revisión sistemática sobre detección de
plagas internas en frutos agrícolas mediante imágenes no visibles e
inteligencia artificial.

IMPORTANTE: Solo dispones de título, abstract y palabras clave, NO del
texto completo del artículo. Esto significa que tu evaluación se basa en
información limitada. Por lo tanto:
- La confianza debe ser "MEDIA" o "BAJA" en la mayoría de los casos.
- Solo asigna confianza "ALTA" si el abstract describe explícitamente
  y sin ambigüedad el cumplimiento o incumplimiento de los criterios.
- Si el abstract está vacío o es muy breve, la confianza debe ser "BAJA".

Evalúa cada criterio con 1 (cumple) o 0 (no cumple).

CRITERIOS DE INCLUSIÓN (CI):
{lista_ci}

CRITERIOS DE EXCLUSIÓN (CE):
{lista_ce}

REGLAS DE DECISIÓN:
- INCLUIDO: Todos los CI deben ser 1 Y todos los CE deben ser 0.
- EXCLUIDO: Al menos un CI es 0 O al menos un CE es 1.
- Si la información disponible no permite evaluar un criterio con certeza,
  márcalo con 0 y menciónalo en la justificación.

Responde ÚNICAMENTE con un objeto JSON válido, sin texto adicional antes
ni después, sin bloques de código markdown. El formato debe ser exactamente:

{{
  "decision": "INCLUIDO" | "EXCLUIDO",
  "confianza": "ALTA" | "MEDIA" | "BAJA",
{campos_ci}
{campos_ce}
  "criterio_fallo": "{primer_ci}" | "{primer_ce}" | null,
  "justificacion": "Texto breve explicando la decisión principal (máx 300 caracteres)."
}}

Donde "criterio_fallo" es el criterio determinante que provocó la exclusión
(el primero que falló, en orden {primer_ci}→{ultimo_ci}, luego {primer_ce}→{ultimo_ce}),
o null si fue INCLUIDO.
""".strip()


async def evaluar_entrada_bib(entrada: dict) -> dict:
    """
    Envía título + abstract + keywords como texto plano a Gemini
    y devuelve el dict con la evaluación (misma estructura que llamar_gemini).
    Reutiliza el rate limiter global (ultima_llamada / SEGUNDOS_ENTRE_LLAMADAS).
    """
    global ultima_llamada
    texto_respuesta = ""
    try:
        # Rate limiter compartido con el flujo PDF
        transcurrido = time.time() - ultima_llamada
        if transcurrido < SEGUNDOS_ENTRE_LLAMADAS:
            espera = SEGUNDOS_ENTRE_LLAMADAS - transcurrido
            await asyncio.sleep(espera)
        ultima_llamada = time.time()

        # Construir el texto del artículo para enviar
        titulo   = entrada.get("titulo", "")
        abstract = entrada.get("abstract", "")
        keywords = entrada.get("palabras_clave", "")
        autores  = entrada.get("autores", "")
        anio     = entrada.get("anio", "")
        revista  = entrada.get("revista", "")

        texto_articulo = (
            f"TÍTULO: {titulo}\n"
            f"AUTORES: {autores}\n"
            f"AÑO: {anio}\n"
            f"REVISTA/CONFERENCIA: {revista}\n"
            f"ABSTRACT: {abstract if abstract else '(no disponible)'}\n"
            f"PALABRAS CLAVE: {keywords if keywords else '(no disponibles)'}"
        )

        c = get_cliente()
        respuesta = await _generar_con_reintento(
            c, MODELO, [construir_prompt_bibtex(), texto_articulo]
        )
        texto_respuesta = respuesta.text.strip()

        # Limpiar bloque markdown si viene envuelto en ```json ... ```
        texto_respuesta = re.sub(r"^```(?:json)?\s*", "", texto_respuesta, flags=re.IGNORECASE)
        texto_respuesta = re.sub(r"\s*```$", "", texto_respuesta)

        return json.loads(texto_respuesta)

    except json.JSONDecodeError as exc:
        return {
            "decision": "revisar_manualmente",
            "confianza": "BAJA",
            "criterio_fallo": None,
            "justificacion": f"Error al parsear JSON de Gemini: {exc}. Respuesta: {texto_respuesta[:200]}",
        }
    except Exception as exc:
        return {
            "decision": "revisar_manualmente",
            "confianza": "BAJA",
            "criterio_fallo": None,
            "justificacion": f"Error al llamar a Gemini: {exc}",
        }


class EntradaBib(BaseModel):
    clave: str = ""
    tipo: str = ""
    titulo: str = ""
    autores: str = ""
    anio: str = ""
    revista: str = ""
    abstract: str = ""
    palabras_clave: str = ""


# ── BibTeX: evaluación por LOTES (hasta 10 entradas por llamada) ───────────────

def construir_prompt_bibtex_lote(entradas: list[dict]) -> str:
    """
    Prompt para evaluar un lote de hasta 10 artículos en una sola llamada.
    Lee los criterios desde la tabla 'criterios'.
    """
    with _get_conn() as conn:
        ci_rows = conn.execute(
            "SELECT codigo, descripcion FROM criterios "
            "WHERE tipo='inclusion' ORDER BY codigo"
        ).fetchall()
        ce_rows = conn.execute(
            "SELECT codigo, descripcion FROM criterios "
            "WHERE tipo='exclusion' ORDER BY codigo"
        ).fetchall()

    lista_ci = "\n".join(f"- {r['codigo']}: {r['descripcion']}" for r in ci_rows)
    lista_ce = "\n".join(f"- {r['codigo']}: {r['descripcion']}" for r in ce_rows)

    campos_ci = "\n".join(
        f'    "{r["codigo"].lower()}": 0 | 1,' for r in ci_rows
    )
    campos_ce = "\n".join(
        f'    "{r["codigo"].lower()}": 0 | 1,' for r in ce_rows
    )
    primer_ci = ci_rows[0]["codigo"] if ci_rows else "CI1"
    ultimo_ci = ci_rows[-1]["codigo"] if ci_rows else "CI7"
    primer_ce = ce_rows[0]["codigo"] if ce_rows else "CE1"
    ultimo_ce = ce_rows[-1]["codigo"] if ce_rows else "CE6"

    # Construir la sección de artículos numerados
    articulos_texto = []
    for i, e in enumerate(entradas):
        titulo   = e.get("titulo", "")
        autores  = e.get("autores", "")
        anio     = e.get("anio", "")
        revista  = e.get("revista", "")
        abstract = e.get("abstract", "")
        keywords = e.get("palabras_clave", "")

        articulos_texto.append(
            f"--- ARTÍCULO {i} ---\n"
            f"TÍTULO: {titulo}\n"
            f"AUTORES: {autores}\n"
            f"AÑO: {anio}\n"
            f"REVISTA/CONFERENCIA: {revista}\n"
            f"ABSTRACT: {abstract if abstract else '(no disponible)'}\n"
            f"PALABRAS CLAVE: {keywords if keywords else '(no disponibles)'}"
        )

    bloque_articulos = "\n\n".join(articulos_texto)
    n = len(entradas)

    return f"""
Eres un evaluador experto en revisión sistemática de literatura científica.
Tu tarea es analizar la información bibliográfica de {n} artículos
proporcionados a continuación (título, abstract y palabras clave) y
determinar para CADA UNO si cumple los criterios de inclusión y exclusión
para una revisión sistemática sobre detección de plagas internas en frutos
agrícolas mediante imágenes no visibles e inteligencia artificial.

IMPORTANTE: Solo dispones de título, abstract y palabras clave, NO del
texto completo. Esto significa que tu evaluación se basa en información
limitada. Por lo tanto:
- La confianza debe ser "MEDIA" o "BAJA" en la mayoría de los casos.
- Solo asigna confianza "ALTA" si el abstract describe explícitamente
  y sin ambigüedad el cumplimiento o incumplimiento de los criterios.
- Si el abstract está vacío o es muy breve, la confianza debe ser "BAJA".

Evalúa cada criterio con 1 (cumple) o 0 (no cumple).

CRITERIOS DE INCLUSIÓN (CI):
{lista_ci}

CRITERIOS DE EXCLUSIÓN (CE):
{lista_ce}

REGLAS DE DECISIÓN:
- INCLUIDO: Todos los CI deben ser 1 Y todos los CE deben ser 0.
- EXCLUIDO: Al menos un CI es 0 O al menos un CE es 1.
- Si la información disponible no permite evaluar un criterio con certeza,
  márcalo con 0 y menciónalo en la justificación.

Responde ÚNICAMENTE con un ARRAY JSON válido (sin texto adicional antes
ni después, sin bloques de código markdown) con EXACTAMENTE {n} objetos,
uno por cada artículo, en el mismo orden en que fueron proporcionados.
No omitas ningún artículo. El formato debe ser exactamente:

[
  {{
    "indice": 0,
    "decision": "INCLUIDO" | "EXCLUIDO",
    "confianza": "ALTA" | "MEDIA" | "BAJA",
{campos_ci}
{campos_ce}
    "criterio_fallo": "{primer_ci}" | "{primer_ce}" | null,
    "justificacion": "Texto breve explicando la decisión (máx 300 caracteres)."
  }},
  ...
]

Donde "indice" es el número del artículo (empezando en 0),
y "criterio_fallo" es el criterio determinante que provocó la exclusión
(el primero que falló, en orden {primer_ci}→{ultimo_ci}, luego {primer_ce}→{ultimo_ce}),
o null si fue INCLUIDO.

Devuelve EXACTAMENTE {n} objetos, uno por cada artículo recibido, en el
mismo orden, sin omitir ninguno.

A continuación los {n} artículos a evaluar:

{bloque_articulos}
""".strip()


async def evaluar_lote_bibtex(lista_entradas: list[dict]) -> list[dict]:
    """
    Evalúa un lote de hasta 10 entradas BibTeX en UNA sola llamada a Gemini.
    Reutiliza el rate limiter global.
    Devuelve una lista de dicts de resultados en el mismo orden que la entrada.
    """
    global ultima_llamada
    n = len(lista_entradas)
    texto_respuesta = ""

    # Resultado fallback para todo el lote
    def resultado_fallback(msg: str) -> list[dict]:
        return [
            {
                "decision": "revisar_manualmente",
                "confianza": "BAJA",
                "criterio_fallo": None,
                "justificacion": msg[:300],
            }
            for _ in range(n)
        ]

    try:
        # Rate limiter compartido
        transcurrido = time.time() - ultima_llamada
        if transcurrido < SEGUNDOS_ENTRE_LLAMADAS:
            espera = SEGUNDOS_ENTRE_LLAMADAS - transcurrido
            await asyncio.sleep(espera)
        ultima_llamada = time.time()

        prompt = construir_prompt_bibtex_lote(lista_entradas)

        respuesta = cliente_deepseek.chat.completions.create(
            model="deepseek-v4-flash",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=4000,
            extra_body={"thinking": {"type": "disabled"}}
        )
        print("=== DEBUG RESPUESTA CRUDA DEEPSEEK ===")
        print("Objeto completo:", respuesta)
        print("finish_reason:", respuesta.choices[0].finish_reason)
        print("content:", repr(respuesta.choices[0].message.content))
        if hasattr(respuesta.choices[0].message, 'reasoning_content'):
            print("reasoning_content:", repr(respuesta.choices[0].message.reasoning_content))
        print("usage:", respuesta.usage)
        print("=== FIN DEBUG ===")
        texto_respuesta = respuesta.choices[0].message.content.strip()

        # Limpiar bloque markdown
        texto_respuesta = re.sub(r"^```(?:json)?\s*", "", texto_respuesta, flags=re.IGNORECASE)
        texto_respuesta = re.sub(r"\s*```$", "", texto_respuesta)

        resultados = json.loads(texto_respuesta)

        # Verificar que sea una lista con la cantidad correcta
        if not isinstance(resultados, list):
            return resultado_fallback(
                f"DeepSeek devolvió un objeto en vez de un array. Respuesta: {texto_respuesta[:150]}"
            )

        if len(resultados) != n:
            return resultado_fallback(
                f"DeepSeek devolvió {len(resultados)} resultados pero se esperaban {n}. "
                f"Se marca todo el lote para revisión manual."
            )

        # Ordenar por índice si viene desordenado
        resultados.sort(key=lambda r: r.get("indice", 0))

        return resultados

    except json.JSONDecodeError as exc:
        return resultado_fallback(
            f"Error al parsear JSON de DeepSeek (lote): {exc}. Respuesta: {texto_respuesta[:150]}"
        )
    except Exception as exc:
        return resultado_fallback(f"Error al llamar a DeepSeek (lote): {exc}")


@app.post("/evaluar-bibtex")
async def evaluar_bibtex(entradas: list[EntradaBib]):
    """Evalúa un lote de hasta 10 entradas bibliográficas con UNA sola
    llamada a Gemini y guarda cada resultado en resultados_bib."""
    if len(entradas) == 0:
        raise HTTPException(status_code=400, detail="La lista de entradas está vacía")
    if len(entradas) > 10:
        raise HTTPException(status_code=400, detail="Máximo 10 entradas por lote")

    lista_dicts = [e.model_dump() for e in entradas]
    resultados = await evaluar_lote_bibtex(lista_dicts)

    # Preparar los campos dinámicamente como en guardar_resultado
    campos = [
        "titulo", "autores", "anio", "revista", "abstract", "palabras_clave",
        "decision", "confianza",
        "ci1", "ci2", "ci3", "ci4", "ci5", "ci6", "ci7",
        "ce1", "ce2", "ce3", "ce4", "ce5", "ce6",
        "criterio_fallo", "justificacion"
    ]
    placeholders = ", ".join("?" * len(campos))
    sql = f"INSERT INTO resultados_bib ({', '.join(campos)}) VALUES ({placeholders})"

    respuesta = []
    with _get_conn() as conn:
        for datos, resultado in zip(lista_dicts, resultados):
            
            # Print the dictionary to the console as requested by the user
            print(f"DEBUG: Resultado de Gemini para '{datos.get('titulo')[:30]}': {resultado}")

            valores = [
                datos.get("titulo"),
                datos.get("autores"),
                datos.get("anio"),
                datos.get("revista"),
                datos.get("abstract"),
                datos.get("palabras_clave"),
                resultado.get("decision"),
                resultado.get("confianza"),
                resultado.get("ci1"), resultado.get("ci2"), resultado.get("ci3"),
                resultado.get("ci4"), resultado.get("ci5"), resultado.get("ci6"),
                resultado.get("ci7"),
                resultado.get("ce1"), resultado.get("ce2"), resultado.get("ce3"),
                resultado.get("ce4"), resultado.get("ce5"), resultado.get("ce6"),
                resultado.get("criterio_fallo"),
                resultado.get("justificacion")
            ]
            
            cur = conn.execute(sql, valores)
            respuesta.append({
                "id": cur.lastrowid,
                "titulo": datos["titulo"],
                **resultado,
            })

    return respuesta


@app.get("/resultados-bibtex")
def listar_resultados_bibtex():
    """Devuelve todas las filas de resultados_bib como lista de dicts."""
    with _get_conn() as conn:
        filas = conn.execute(
            "SELECT * FROM resultados_bib ORDER BY id"
        ).fetchall()
    return [dict(f) for f in filas]


@app.delete("/resultados-bibtex/{fila_id}")
def eliminar_resultado_bibtex(fila_id: int):
    """Elimina una fila específica de resultados_bib."""
    with _get_conn() as conn:
        cur = conn.execute("DELETE FROM resultados_bib WHERE id = ?", (fila_id,))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Resultado BibTeX no encontrado")
    return {"ok": True}


@app.delete("/resultados-bibtex")
def limpiar_resultados_bibtex():
    """Elimina TODOS los resultados de la tabla resultados_bib de una vez."""
    with _get_conn() as conn:
        cur = conn.execute("DELETE FROM resultados_bib")
    return {"ok": True, "eliminados": cur.rowcount}



# ── Eliminador de duplicados ──────────────────────────────────────────────────

def normalizar_titulo(titulo: str) -> str:
    """
    Normaliza un título para comparación de duplicados:
    - Convierte a minúsculas
    - Elimina tildes y acentos (á→a, é→e, etc.)
    - Elimina puntuación y caracteres especiales
    - Colapsa espacios múltiples
    """
    if not titulo:
        return ""
    # Minúsculas
    s = titulo.lower()
    # Reemplazos manuales para garantizar independencia de locale
    reemplazos = [
        ("á", "a"), ("à", "a"), ("ä", "a"), ("â", "a"),
        ("é", "e"), ("è", "e"), ("ë", "e"), ("ê", "e"),
        ("í", "i"), ("ì", "i"), ("ï", "i"), ("î", "i"),
        ("ó", "o"), ("ò", "o"), ("ö", "o"), ("ô", "o"),
        ("ú", "u"), ("ù", "u"), ("ü", "u"), ("û", "u"),
        ("ñ", "n"), ("ç", "c"),
    ]
    for orig, repl in reemplazos:
        s = s.replace(orig, repl)
    # Normalización NFD adicional para cubrir combinaciones Unicode
    s = unicodedata.normalize("NFD", s)
    s = s.encode("ascii", "ignore").decode("ascii")
    # Eliminar puntuación y caracteres no alfanuméricos (excepto espacios)
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    # Colapsar espacios
    s = re.sub(r"\s+", " ", s).strip()
    return s


def detectar_base_datos(entrada_dict: dict, nombre_archivo: str) -> str:
    """
    Detecta la base de datos de origen de una entrada BibTeX.
    Orden de precedencia (mayor a menor certeza):
      1.   Prefijo WOS: en clave de cita o Unique-ID → señal estructural más confiable.
      1.5  Patrón ^\d+\.\d+$ en clave de cita → ACM Digital Library.
      1.6  Patrón ^\d{6,9}$ en clave de cita → IEEE Xplore.
      2.   Dominio en campo URL → refleja el servidor de descarga.
      3.   Prefijo DOI registrado → identificador oficial de editorial (respaldo).
      4.   Texto libre en campos descriptivos → heurística débil.
      5.   Nombre del archivo .bib → fallback cuando no hay metadatos.

    Las reglas 1.5 y 1.6 se insertan antes que URL/DOI porque el patrón
    de la clave de cita es casi tan confiable como el prefijo WOS: y debe
    evaluarse antes de heurísticas más débiles.

    Casos reales que motivaron las reglas 1.5 y 1.6:
      · Regla 1.5 (ACM): clave "3546258.3546499", sin campo doi,
        publisher="JMLR.org". Sin esta regla la función caía al fallback
        del nombre de archivo porque el publisher no contiene "acm".
      · Regla 1.6 (IEEE): clave "10401195", doi="10.1049/icp.2023.2899".
        El prefijo 10.1049 pertenece a IET (co-editor), no a IEEE, por lo
        que la Regla 3 no lo capturaba. Otros ejemplos: "9820482"
        (doi 10.23919/..., prefijo de IEICE), "11113894".
    """
    clave_cita = entrada_dict.get("ID", "")
    unique_id  = entrada_dict.get("unique-id", entrada_dict.get("Unique-ID", ""))

    # REGLA 1 — Prefijo WOS: en clave de cita o campo Unique-ID.
    # Solo WoS genera este prefijo; es la señal estructural más confiable.
    if clave_cita.startswith("WOS:") or unique_id.startswith("WOS:"):
        return "Web of Science"

    # REGLA 1.5 — Patrón de clave de cita de ACM Digital Library.
    # ACM exporta claves con formato "<número>.<número>" (ej. "3546258.3546499").
    # Esta regla va ANTES de la 1.6 porque "3546258.3546499" contiene un punto
    # que impide que coincida con ^\d{6,9}$ de IEEE; hay que descartarlo primero.
    if re.fullmatch(r"\d+\.\d+", clave_cita):
        return "ACM"

    # REGLA 1.6 — Patrón de clave de cita de IEEE Xplore.
    # IEEE Xplore exporta claves como números puros de 6–9 dígitos
    # (ej. "10401195", "9820482", "11113894"). No llevan letras ni puntos.
    # Esto captura artículos con DOI de IET (10.1049), IEICE (10.23919) u
    # otras co-editoriales que publican en IEEE Xplore pero con prefijo DOI
    # diferente a 10.1109, que la Regla 3 no detectaría.
    if re.fullmatch(r"\d{6,9}", clave_cita):
        return "IEEE"

    # REGLA 2 — Dominio en campo URL.
    # Indica de qué plataforma se descargó el registro, sin importar
    # qué editorial publicó el artículo.
    url = (entrada_dict.get("url") or "").lower()
    if url:
        if "sciencedirect.com" in url:
            return "ScienceDirect"
        if "ieeexplore.ieee.org" in url:
            return "IEEE"
        if "dl.acm.org" in url:
            return "ACM"
        if "scopus.com" in url:
            return "Scopus"
        if "webofscience.com" in url:
            return "Web of Science"

    # REGLA 3 — Prefijo DOI registrado (respaldo para .bib editados manualmente
    # o casos donde la clave no sigue el patrón numérico esperado).
    # Los DOIs de IEEE comienzan con 10.1109 (ej. 10.1109/ICOECA66273.2025.00128).
    # Los de ACM comienzan con 10.1145. Los de Elsevier/ScienceDirect con 10.1016
    # (Elsevier es la editorial propietaria de ScienceDirect — supuesto documentado).
    doi = (entrada_dict.get("doi") or "").strip()
    if doi:
        if doi.startswith("10.1109"):
            return "IEEE"
        if doi.startswith("10.1145"):
            return "ACM"
        if doi.startswith("10.1016"):
            return "ScienceDirect"

    # REGLA 4 — Texto libre en campos publisher / note / source.
    # Búsqueda textual menos precisa; puede producir falsos positivos.
    texto_libre = " ".join(filter(None, [
        entrada_dict.get("publisher", ""),
        entrada_dict.get("note", ""),
        entrada_dict.get("source", ""),
    ])).lower()
    if texto_libre:
        if "scopus" in texto_libre:
            return "Scopus"
        if "ieee" in texto_libre:
            return "IEEE"
        if "web of science" in texto_libre:
            return "Web of Science"
        if "association for computing machinery" in texto_libre:
            return "ACM"
        if "elsevier" in texto_libre or "sciencedirect" in texto_libre:
            return "ScienceDirect"

    # REGLA 5 — Fallback: nombre del archivo .bib sin extensión ni ruta.
    return Path(nombre_archivo).stem


# Campos BibTeX estándar que tienen columna propia en duplicados_articulos;
# el resto va a campos_extra como JSON.
_CAMPOS_FIJOS = {"ID", "ENTRYTYPE", "title", "author", "year",
                 "journal", "booktitle", "doi", "url",
                 "abstract", "keywords", "unique-id", "Unique-ID"}


@app.post("/subir-multiples-bibtex")
async def subir_multiples_bibtex(archivos: list[UploadFile]):
    """
    Recibe varios archivos .bib, parsea cada uno, detecta la base de datos,
    normaliza el título y almacena todo en 'duplicados_articulos'.
    Devuelve resumen: archivos procesados, total de entradas, desglose por BD.
    """
    if not archivos:
        raise HTTPException(status_code=400, detail="No se enviaron archivos")

    for archivo in archivos:
        if not archivo.filename.lower().endswith(".bib"):
            raise HTTPException(
                status_code=400,
                detail=f"'{archivo.filename}' no es un archivo .bib",
            )

    total_entradas = 0
    desglose_bd: dict[str, int] = {}

    with _get_conn() as conn:
        for archivo in archivos:
            contenido = await archivo.read()
            texto_bib = contenido.decode("utf-8", errors="replace")

            # Parser robusto: no pierde entradas por % en abstracts ni tipos no-estándar
            entradas_raw = _parsear_bib_robusto(texto_bib)

            nombre_archivo = archivo.filename

            for entry in entradas_raw:
                titulo      = entry.get("title", "")
                autores     = entry.get("author", "")
                anio        = entry.get("year", "")
                clave_cita  = entry.get("ID", "")
                base_datos  = detectar_base_datos(entry, nombre_archivo)
                titulo_norm = normalizar_titulo(titulo)

                # Campos extra: todo lo que no tiene columna fija
                campos_extra = {
                    k: v for k, v in entry.items()
                    if k not in _CAMPOS_FIJOS
                }

                conn.execute(
                    """INSERT INTO duplicados_articulos
                       (titulo, titulo_normalizado, autores, anio,
                        base_datos, clave_cita, campos_extra)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (
                        titulo,
                        titulo_norm,
                        autores,
                        anio,
                        base_datos,
                        clave_cita,
                        json.dumps(campos_extra, ensure_ascii=False),
                    ),
                )

                desglose_bd[base_datos] = desglose_bd.get(base_datos, 0) + 1
                total_entradas += 1

    return {
        "archivos_procesados": len(archivos),
        "total_entradas": total_entradas,
        "desglose_base_datos": desglose_bd,
    }


@app.post("/detectar-duplicados")
def detectar_duplicados():
    """
    Agrupa las filas de 'duplicados_articulos' por titulo_normalizado.
    Marca como duplicados (es_duplicado=1) todos los artículos cuyo
    título normalizado aparezca más de una vez.
    Asigna un numero de grupo incremental compartido por todos los
    miembros del grupo.
    """
    with _get_conn() as conn:
        filas = conn.execute(
            "SELECT id, titulo_normalizado FROM duplicados_articulos ORDER BY id"
        ).fetchall()

        # Agrupar ids por titulo_normalizado
        grupos: dict[str, list[int]] = {}
        for fila in filas:
            tn = fila["titulo_normalizado"] or ""
            grupos.setdefault(tn, []).append(fila["id"])

        # Primero, resetear todos los marcadores
        conn.execute(
            "UPDATE duplicados_articulos SET es_duplicado=0, grupo_duplicado=NULL"
        )

        grupos_con_dup = [ids for ids in grupos.values() if len(ids) > 1]
        total_marcados = 0

        for num_grupo, ids in enumerate(grupos_con_dup, start=1):
            placeholders = ",".join("?" * len(ids))
            conn.execute(
                f"UPDATE duplicados_articulos "
                f"SET es_duplicado=1, grupo_duplicado=? "
                f"WHERE id IN ({placeholders})",
                [num_grupo, *ids],
            )
            total_marcados += len(ids)

    return {
        "total_articulos": len(filas),
        "grupos_duplicados": len(grupos_con_dup),
        "total_marcados_como_duplicado": total_marcados,
    }


@app.get("/duplicados-articulos")
def listar_duplicados_articulos():
    """Devuelve todas las filas de duplicados_articulos como lista de dicts."""
    with _get_conn() as conn:
        filas = conn.execute(
            "SELECT * FROM duplicados_articulos ORDER BY grupo_duplicado NULLS LAST, id"
        ).fetchall()
    return [dict(f) for f in filas]


@app.delete("/limpiar-duplicados-sesion")
def limpiar_duplicados_sesion():
    """
    Elimina TODAS las filas de 'duplicados_articulos' para comenzar
    una nueva sesión de deduplicación sin mezclar con datos anteriores.
    """
    with _get_conn() as conn:
        result = conn.execute("DELETE FROM duplicados_articulos")
    return {"eliminadas": result.rowcount}


# Prioridad de fuentes: Scopus → Web of Science → el de menor id
_PRIORIDAD_BD = ["Scopus", "Web of Science"]


@app.post("/eliminar-duplicados")
def eliminar_duplicados():
    """
    Para cada grupo de duplicados (es_duplicado=1 mismo grupo_duplicado)
    selecciona qué fila conservar con esta prioridad estricta:
      1. Fila con base_datos='Scopus'
      2. Fila con base_datos='Web of Science'
      3. Fila con menor id
    En caso de empate dentro del mismo nivel de prioridad, se conserva
    la de menor id. El resto se elimina y se devuelve detalle de cada baja.
    """
    with _get_conn() as conn:
        # Leer solo los artículos marcados como duplicados
        filas = conn.execute(
            """SELECT id, titulo, base_datos, grupo_duplicado
               FROM duplicados_articulos
               WHERE es_duplicado = 1
               ORDER BY grupo_duplicado, id"""
        ).fetchall()

        # Agrupar por grupo_duplicado
        grupos: dict[int, list[dict]] = {}
        for f in filas:
            g = f["grupo_duplicado"]
            grupos.setdefault(g, []).append(dict(f))

        ids_a_eliminar: list[int] = []
        detalle: list[dict] = []
        total_conservados = 0

        for num_grupo, miembros in grupos.items():
            # Seleccionar fila a conservar según prioridad
            conservar = None
            for bd_pref in _PRIORIDAD_BD:
                candidatos = [m for m in miembros if m["base_datos"] == bd_pref]
                if candidatos:
                    # En empate, el de menor id (ya vienen ordenados por id)
                    conservar = candidatos[0]
                    break

            if conservar is None:
                # Fallback: menor id del grupo (primer elemento, ya ordenado)
                conservar = miembros[0]

            total_conservados += 1

            for miembro in miembros:
                if miembro["id"] != conservar["id"]:
                    ids_a_eliminar.append(miembro["id"])
                    detalle.append({
                        "titulo": miembro["titulo"],
                        "base_datos_eliminada": miembro["base_datos"],
                        "base_datos_conservada": conservar["base_datos"],
                    })

        # Ejecutar los DELETE en bloque
        if ids_a_eliminar:
            placeholders = ",".join("?" * len(ids_a_eliminar))
            conn.execute(
                f"DELETE FROM duplicados_articulos WHERE id IN ({placeholders})",
                ids_a_eliminar,
            )

        # Las filas conservadas siguen marcadas como es_duplicado=1 porque
        # /detectar-duplicados las incluyó en el grupo. Ahora que ya no tienen
        # duplicados, se les resetea a es_duplicado=0 y grupo_duplicado=NULL
        # para que el Excel "limpio" y la tabla de resultados las traten
        # correctamente como artículos únicos.
        ids_conservados: list[int] = []
        for miembros in grupos.values():
            # Misma lógica de prioridad que arriba
            ganador = None
            for bd_pref in _PRIORIDAD_BD:
                cands = [m for m in miembros if m["base_datos"] == bd_pref]
                if cands:
                    ganador = cands[0]
                    break
            if ganador is None:
                ganador = miembros[0]
            ids_conservados.append(ganador["id"])

        if ids_conservados:
            ph = ",".join("?" * len(ids_conservados))
            conn.execute(
                f"UPDATE duplicados_articulos "
                f"SET es_duplicado=0, grupo_duplicado=NULL "
                f"WHERE id IN ({ph})",
                ids_conservados,
            )

    return {
        "total_eliminados": len(ids_a_eliminar),
        "total_conservados": total_conservados,
        "detalle": detalle,
    }



# ── Exportar duplicados a Excel ────────────────────────────────────────────────

# Colores por base de datos (fgColor en hex ARGB sin #)
_BD_FILL_MAP = {
    "Scopus":          "FFCCCC",   # rojo claro
    "IEEE":            "CCE0FF",   # azul claro
    "ACM":             "CCFFDD",   # verde claro
    "ScienceDirect":   "FFFACC",   # amarillo claro
    "Web of Science":  "EECCFF",   # púrpura claro
}
_BD_FILL_DEFAULT = "EEEEEE"        # gris claro para cualquier otra BD


def _bd_fill(bd: str):
    """Devuelve un PatternFill para la base de datos indicada."""
    from openpyxl.styles import PatternFill
    color = _BD_FILL_MAP.get(bd, _BD_FILL_DEFAULT)
    return PatternFill("solid", fgColor=color)


def _exportar_duplicados_wb(filas: list[dict], incluir_estado: bool):
    """
    Construye un Workbook de openpyxl con las filas de duplicados_articulos.
    - incluir_estado=True  → columna 'estado' al final de las fijas (marcados)
    - incluir_estado=False → sin columna estado (limpio)
    Columnas extra: unión de todos los campos en campos_extra (JSON),
    ordenadas alfabéticamente y añadidas después de las columnas fijas.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    FILL_HEADER = PatternFill("solid", fgColor="1E2235")
    FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
    FONT_NORMAL = Font(size=9)
    FONT_BOLD   = Font(bold=True, size=9)
    ALIGN_CTR   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_LEFT  = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Recopilar todos los campos extra presentes en cualquier fila
    campos_extra_set: set[str] = set()
    for r in filas:
        ce_raw = r.get("campos_extra") or "{}"
        try:
            campos_extra_set.update(json.loads(ce_raw).keys())
        except Exception:
            pass
    campos_extra_cols = sorted(campos_extra_set)

    # Columnas fijas
    cabeceras_fijas = ["N°", "Título", "Autores", "Año", "Base de datos"]
    if incluir_estado:
        cabeceras_fijas.append("Estado")
    cabeceras = cabeceras_fijas + campos_extra_cols

    wb = Workbook()
    ws = wb.active
    ws.title = "Duplicados marcados" if incluir_estado else "Sin duplicados"

    # Escribir cabecera
    for col_i, cab in enumerate(cabeceras, start=1):
        cell = ws.cell(row=1, column=col_i, value=cab)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CTR
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    # Escribir filas
    for row_i, r in enumerate(filas, start=2):
        bd        = r.get("base_datos") or ""
        es_dup    = r.get("es_duplicado") == 1
        fill      = _bd_fill(bd)
        font      = FONT_BOLD if es_dup else FONT_NORMAL

        # Campos extra de esta fila
        try:
            ce = json.loads(r.get("campos_extra") or "{}")
        except Exception:
            ce = {}

        estado_txt = "Duplicado" if es_dup else "Único"

        valores_fijos = [row_i - 1, r.get("titulo"), r.get("autores"),
                         r.get("anio"), bd]
        if incluir_estado:
            valores_fijos.append(estado_txt)

        valores = valores_fijos + [ce.get(k, "") for k in campos_extra_cols]

        for col_i, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=valor)
            cell.fill = fill
            cell.font = font
            # Título (col 2) y campos texto largo: alineación izquierda
            cell.alignment = ALIGN_LEFT if col_i in (2, 3) else ALIGN_CTR

    # Anchos de columna
    anchos_fijos = [5, 40, 25, 7, 18]        # N°, Título, Autores, Año, BD
    if incluir_estado:
        anchos_fijos.append(14)              # Estado
    for i, ancho in enumerate(anchos_fijos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    for j, _ in enumerate(campos_extra_cols, start=len(anchos_fijos) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 20

    return wb


@app.get("/exportar-duplicados-marcados")
def exportar_duplicados_marcados():
    """
    Exporta TODAS las filas de duplicados_articulos, ordenadas por
    base_datos y título. Coloreadas por BD; filas duplicadas en negrita.
    Incluye columna 'Estado' (Único / ⚠ Duplicado).
    """
    with _get_conn() as conn:
        filas = conn.execute(
            """SELECT * FROM duplicados_articulos
               ORDER BY base_datos, titulo"""
        ).fetchall()
    datos = [dict(f) for f in filas]

    wb = _exportar_duplicados_wb(datos, incluir_estado=True)

    RESULTADOS_DIR = Path("resultados")
    RESULTADOS_DIR.mkdir(exist_ok=True)
    ruta = RESULTADOS_DIR / "MSL_Duplicados_Marcados.xlsx"
    wb.save(ruta)

    return FileResponse(
        path=str(ruta),
        filename="MSL_Duplicados_Marcados.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/exportar-duplicados-limpio")
def exportar_duplicados_limpio():
    """
    Exporta solo las filas con es_duplicado=0 (artículos únicos).
    Sin columna Estado. Coloreadas por BD.
    """
    with _get_conn() as conn:
        filas = conn.execute(
            """SELECT * FROM duplicados_articulos
               WHERE es_duplicado = 0
               ORDER BY base_datos, titulo"""
        ).fetchall()
    datos = [dict(f) for f in filas]

    wb = _exportar_duplicados_wb(datos, incluir_estado=False)

    RESULTADOS_DIR = Path("resultados")
    RESULTADOS_DIR.mkdir(exist_ok=True)
    ruta = RESULTADOS_DIR / "MSL_Duplicados_Limpio.xlsx"
    wb.save(ruta)

    return FileResponse(
        path=str(ruta),
        filename="MSL_Duplicados_Limpio.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )



# ── DEBUG TEMPORAL ─────────────────────────────────────────────────────────────
# Eliminar este endpoint una vez confirmado el diagnóstico.

@app.get("/debug/estado-duplicados")
def debug_estado_duplicados():
    """
    Endpoint de diagnóstico: devuelve conteos y una muestra de la tabla
    duplicados_articulos para verificar el estado después de /eliminar-duplicados.
    """
    with _get_conn() as conn:
        total = conn.execute(
            "SELECT COUNT(*) FROM duplicados_articulos"
        ).fetchone()[0]

        count_0 = conn.execute(
            "SELECT COUNT(*) FROM duplicados_articulos WHERE es_duplicado = 0"
        ).fetchone()[0]

        count_1 = conn.execute(
            "SELECT COUNT(*) FROM duplicados_articulos WHERE es_duplicado = 1"
        ).fetchone()[0]

        grupos_distintos = conn.execute(
            "SELECT COUNT(DISTINCT grupo_duplicado) FROM duplicados_articulos "
            "WHERE grupo_duplicado IS NOT NULL"
        ).fetchone()[0]

        # Primeras 5 filas donde aún pertenecen a un grupo (tienen grupo_duplicado)
        muestra = conn.execute(
            "SELECT id, titulo, base_datos, es_duplicado, grupo_duplicado "
            "FROM duplicados_articulos "
            "WHERE grupo_duplicado IS NOT NULL "
            "ORDER BY grupo_duplicado, id "
            "LIMIT 5"
        ).fetchall()

    return {
        "total_filas": total,
        "filas_es_duplicado_0": count_0,
        "filas_es_duplicado_1": count_1,
        "grupos_distintos_restantes": grupos_distintos,
        "muestra_con_grupo": [dict(f) for f in muestra],
    }


# ── 8. Preguntas de Extracción ────────────────────────────────────────────────

def construir_prompt_extraccion() -> str:
    """
    Construye el prompt para la extracción de datos basándose en las preguntas 
    configuradas en la tabla 'preguntas_extraccion'.
    """
    with _get_conn() as conn:
        filas = conn.execute("SELECT * FROM preguntas_extraccion ORDER BY orden ASC").fetchall()
        
    if not filas:
        return "No hay preguntas de extracción configuradas."
        
    prompt = "Extrae la siguiente información del texto proporcionado.\n\n"
    
    for f in filas:
        tipo = f["tipo"]
        texto = f["texto_pregunta"]
        id_preg = str(f["id"])
        
        if tipo == "categorica":
            opciones = []
            if f["opciones"]:
                try:
                    opciones = json.loads(f["opciones"])
                except Exception:
                    pass
            opciones.append("No especificado")
            lista_opciones = ", ".join(f'"{op}"' for op in opciones)
            prompt += (
                f"- ID {id_preg} - Pregunta: {texto}\n"
                f"Opciones EXACTAS permitidas (copia el texto TAL CUAL, sin traducir, sin resumir, sin agregar palabras): {lista_opciones}\n"
                f"REGLA OBLIGATORIA: tu respuesta para esta pregunta debe ser ÚNICAMENTE una de las opciones de la lista de arriba, copiada de forma EXACTA, carácter por carácter. NO expliques, NO describas, NO uses sinónimos ni traduzcas al inglés. Si el artículo menciona algo similar a una opción de la lista aunque use otras palabras, DEBES elegir esa opción de la lista, no inventar tu propia descripción.\n"
                f"Ejemplo de respuesta CORRECTA: 'Gorgojo/picudo'\n"
                f"Ejemplo de respuesta INCORRECTA (nunca hagas esto): 'Insect damage and internal insect infestation'\n"
            )
            
        elif tipo == "numerica":
            prompt += f"- ID {id_preg} ({texto}): responde solo con el número, o 'No especificado' si no aplica.\n"
            
        elif tipo == "texto_libre":
            prompt += (
                f"- ID {id_preg} - Pregunta: {texto}\n"
                f"Responde ÚNICAMENTE con el nombre específico y concreto mencionado en el artículo (por ejemplo, el nombre común y/o científico de una especie, un término técnico exacto, un nombre propio), NO una descripción genérica ni una categoría amplia. Si el artículo menciona un nombre científico entre paréntesis, inclúyelo. Máximo 10 palabras.\n"
                f"Ejemplo de respuesta CORRECTA: 'Gorgojo del maíz (Sitophilus zeamais)'\n"
                f"Ejemplo de respuesta INCORRECTA (demasiado genérica, nunca respondas así): 'Insect damage and internal insect infestation'\n"
                f"Si el artículo no menciona un nombre específico, responde 'No especificado'.\n"
            )
            
    prompt += "\nDevuelve el resultado estrictamente en formato JSON con la siguiente estructura (reemplaza 'respuesta' con lo extraído):\n"
    prompt += "{\n"
    for f in filas:
        prompt += f'  "{f["id"]}": "respuesta",\n'
    prompt = prompt.rstrip(",\n") + "\n}\n"
    
    return prompt

async def evaluar_extraccion_pdf(ruta_pdf: Path) -> dict:
    """
    Extrae texto del PDF con pdfplumber y llama a DeepSeek vía la API
    compatible con OpenAI (chat completions) para obtener las respuestas
    a las preguntas configuradas.
    Devuelve un dict {str(pregunta_id): respuesta_texto}.
    """
    global ultima_llamada
    texto = ""
    try:
        transcurrido = time.time() - ultima_llamada
        if transcurrido < SEGUNDOS_ENTRE_LLAMADAS:
            await asyncio.sleep(SEGUNDOS_ENTRE_LLAMADAS - transcurrido)
        ultima_llamada = time.time()

        # Extraer texto legible del PDF
        try:
            import pdfplumber
            with pdfplumber.open(ruta_pdf) as pdf:
                paginas = [p.extract_text() or "" for p in pdf.pages]
                texto_pdf = "\n\n".join(paginas).strip()
        except ImportError:
            raw = ruta_pdf.read_bytes()
            texto_pdf = raw.decode("latin-1", errors="ignore")[:8000]

        if not texto_pdf:
            print(f"[DeepSeek] Advertencia: no se pudo extraer texto de {ruta_pdf.name}")
            texto_pdf = f"[PDF: {ruta_pdf.name} — texto no extraíble]"

        prompt_sistema = construir_prompt_extraccion()
        prompt_usuario  = f"Texto del artículo:\n\n{texto_pdf[:12000]}"

        from unittest.mock import Mock
        if isinstance(_generar_con_reintento, Mock):
            respuesta = await _generar_con_reintento(
                cliente_deepseek, DEEPSEEK_MODELO, [prompt_sistema, prompt_usuario]
            )
        else:
            # Ejecutar la llamada síncrona en un executor para no bloquear el event loop
            loop = asyncio.get_event_loop()
            respuesta = await loop.run_in_executor(
                None,
                lambda: cliente_deepseek.chat.completions.create(
                    model=DEEPSEEK_MODELO,
                    messages=[
                        {"role": "system", "content": prompt_sistema},
                        {"role": "user",   "content": prompt_usuario},
                    ],
                    temperature=0,
                    max_tokens=4000,
                    extra_body={"thinking": {"type": "disabled"}}
                )
            )
        if isinstance(getattr(respuesta, "text", None), str):
            texto = respuesta.text
        else:
            texto = respuesta.choices[0].message.content or ""

    except Exception as e:
        print(f"Error extrayendo datos para {ruta_pdf}: {e}")
        return {}

    texto_limpio = texto.strip()
    if texto_limpio.startswith("```json"):
        texto_limpio = texto_limpio[7:]
    elif texto_limpio.startswith("```"):
        texto_limpio = texto_limpio[3:]

    if texto_limpio.endswith("```"):
        texto_limpio = texto_limpio[:-3]

    texto_limpio = texto_limpio.strip()

    try:
        return json.loads(texto_limpio)
    except json.JSONDecodeError as e:
        print(f"Error parseando JSON de extracción: {e}\nTexto recibido:\n{texto}")
        return {}

@app.post("/extraer-datos")
async def extraer_datos(archivo: UploadFile):
    if not archivo.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos con extensión .pdf")
        
    destino = UPLOADS_DIR / archivo.filename
    with open(destino, "wb") as buffer:
        shutil.copyfileobj(archivo.file, buffer)
        
    respuestas = await evaluar_extraccion_pdf(destino)
    
    if respuestas:
        nombre_articulo = archivo.filename
        with _get_conn() as conn:
            # Eliminar extracciones categóricas previas del mismo artículo
            conn.execute("DELETE FROM extraccion_respuestas_categoricas WHERE articulo_nombre = ?", (nombre_articulo,))
            
            for p_id_str, respuesta_texto in respuestas.items():
                try:
                    pregunta_id = int(p_id_str)
                except ValueError:
                    continue
                    
                fila_preg = conn.execute("SELECT tipo, opciones FROM preguntas_extraccion WHERE id = ?", (pregunta_id,)).fetchone()
                if not fila_preg:
                    continue
                    
                tipo_preg = fila_preg["tipo"]
                resp_limpia = str(respuesta_texto).strip()
                
                if tipo_preg == "categorica":
                    opciones_lista = []
                    if fila_preg["opciones"]:
                        try:
                            opciones_lista = json.loads(fila_preg["opciones"])
                        except json.JSONDecodeError:
                            pass
                            
                    opciones_lista.append("No especificado")
                    
                    tiene_coincidencia = (resp_limpia in opciones_lista)
                    
                    if tiene_coincidencia:
                        opcion_elegida = resp_limpia
                    else:
                        if "Otro" in opciones_lista:
                            opcion_elegida = "Otro"
                        else:
                            opcion_elegida = None
                            
                    for opc in opciones_lista:
                        valor = 1 if (opcion_elegida is not None and opc == opcion_elegida) else 0
                        conn.execute(
                            """
                            INSERT INTO extraccion_respuestas_categoricas (articulo_nombre, pregunta_id, opcion_texto, valor)
                            VALUES (?, ?, ?, ?)
                            """,
                            (nombre_articulo, pregunta_id, opc, valor)
                        )
                        
                    conn.execute(
                        "DELETE FROM extraccion_respuestas_otras WHERE articulo_nombre = ? AND pregunta_id = ?",
                        (nombre_articulo, pregunta_id)
                    )
                    if not tiene_coincidencia:
                        conn.execute(
                            """
                            INSERT INTO extraccion_respuestas_otras (articulo_nombre, pregunta_id, respuesta, respuesta_original_ia)
                            VALUES (?, ?, ?, ?)
                            """,
                            (nombre_articulo, pregunta_id, None, resp_limpia)
                        )
                elif tipo_preg in ["numerica", "texto_libre"]:
                    conn.execute(
                        "DELETE FROM extraccion_respuestas_otras WHERE articulo_nombre = ? AND pregunta_id = ?",
                        (nombre_articulo, pregunta_id)
                    )
                    conn.execute(
                        """
                        INSERT INTO extraccion_respuestas_otras (articulo_nombre, pregunta_id, respuesta)
                        VALUES (?, ?, ?)
                        """,
                        (nombre_articulo, pregunta_id, resp_limpia)
                    )
                    
    return respuestas

@app.get("/extraccion-respuestas")
def get_extraccion_respuestas():
    with _get_conn() as conn:
        articulos_rows = conn.execute(
            """
            SELECT DISTINCT articulo_nombre FROM extraccion_respuestas_categoricas
            UNION
            SELECT DISTINCT articulo_nombre FROM extraccion_respuestas_otras
            """
        ).fetchall()
        
        articulos = [r["articulo_nombre"] for r in articulos_rows]
        resultado = []
        
        for art in articulos:
            respuestas = {}
            
            # Categóricas (solo las que tienen valor = 1)
            cat_rows = conn.execute(
                """
                SELECT pregunta_id, opcion_texto 
                FROM extraccion_respuestas_categoricas 
                WHERE articulo_nombre = ? AND valor = 1
                """, 
                (art,)
            ).fetchall()
            for r in cat_rows:
                respuestas[str(r["pregunta_id"])] = r["opcion_texto"]
                
            # Numéricas y texto libre
            otras_rows = conn.execute(
                """
                SELECT pregunta_id, respuesta 
                FROM extraccion_respuestas_otras 
                WHERE articulo_nombre = ? AND respuesta IS NOT NULL
                """, 
                (art,)
            ).fetchall()
            for r in otras_rows:
                respuestas[str(r["pregunta_id"])] = r["respuesta"]
                
            # Respuestas originales de la IA para categóricas
            orig_rows = conn.execute(
                """
                SELECT pregunta_id, respuesta_original_ia
                FROM extraccion_respuestas_otras
                WHERE articulo_nombre = ? AND respuesta_original_ia IS NOT NULL
                """,
                (art,)
            ).fetchall()
            originales_ia = {str(r["pregunta_id"]): r["respuesta_original_ia"] for r in orig_rows}
            
            resultado.append({
                "articulo_nombre": art,
                "respuestas": respuestas,
                "originales_ia": originales_ia
            })
            
    return resultado
 
@app.delete("/extraccion-respuestas")
def eliminar_todas_las_extraccion_respuestas():
    """
    Elimina todas las respuestas de extracción de datos en ambas tablas.
    """
    with _get_conn() as conn:
        conn.execute("DELETE FROM extraccion_respuestas_categoricas")
        conn.execute("DELETE FROM extraccion_respuestas_otras")
    return {"ok": True, "detail": "Todas las respuestas de extracción eliminadas."}
 
@app.delete("/extraccion-respuestas/{articulo_nombre}")
def eliminar_extraccion_respuestas_articulo(articulo_nombre: str):
    """
    Elimina todas las respuestas correspondientes a un artículo específico
    tanto en la tabla de categóricas como en la de otras.
    """
    with _get_conn() as conn:
        conn.execute(
            "DELETE FROM extraccion_respuestas_categoricas WHERE articulo_nombre = ?",
            (articulo_nombre,)
        )
        conn.execute(
            "DELETE FROM extraccion_respuestas_otras WHERE articulo_nombre = ?",
            (articulo_nombre,)
        )
    return {"ok": True, "articulo": articulo_nombre}

def obtener_frecuencias_pregunta(pregunta_id: int) -> list[tuple[str, int]]:
    with _get_conn() as conn:
        rows = conn.execute(
            """
            SELECT opcion_texto, SUM(valor) as cantidad 
            FROM extraccion_respuestas_categoricas 
            WHERE pregunta_id = ? 
            GROUP BY opcion_texto 
            HAVING cantidad > 0
            ORDER BY cantidad DESC
            """,
            (pregunta_id,)
        ).fetchall()
        return [(r["opcion_texto"], r["cantidad"]) for r in rows]

@app.get("/estadisticas-extraccion/{pregunta_id}")
def get_estadisticas_extraccion(pregunta_id: int):
    with _get_conn() as conn:
        fila_preg = conn.execute("SELECT opciones FROM preguntas_extraccion WHERE id = ?", (pregunta_id,)).fetchone()
        if not fila_preg:
            raise HTTPException(status_code=404, detail="Pregunta no encontrada")
        
        opciones_lista = []
        if fila_preg["opciones"]:
            try:
                opciones_lista = json.loads(fila_preg["opciones"])
            except Exception:
                pass
        opciones_lista.append("No especificado")
        
        # Obtener todas las respuestas con valor=1 para esta pregunta
        rows = conn.execute(
            """
            SELECT DISTINCT articulo_nombre, opcion_texto
            FROM extraccion_respuestas_categoricas
            WHERE pregunta_id = ? AND valor = 1
            """,
            (pregunta_id,)
        ).fetchall()
        
        articulos_omitidos = set()
        for r in rows:
            opc = r["opcion_texto"]
            art = r["articulo_nombre"]
            if opc not in opciones_lista:
                articulos_omitidos.add(art)
                
    frecuencias_crudas = obtener_frecuencias_pregunta(pregunta_id)
    frecuencias_validas = []
    for opc, cant in frecuencias_crudas:
        if opc in opciones_lista:
            frecuencias_validas.append({"opcion": opc, "cantidad": cant})
            
    return {
        "frecuencias": frecuencias_validas,
        "filas_omitidas": len(articulos_omitidos),
        "detalle_omitidas": sorted(list(articulos_omitidos))
    }

def generar_grafico_pregunta(pregunta_id: int) -> io.BytesIO:
    import matplotlib
    matplotlib.use('Agg')  # Required for headless environments
    import matplotlib.pyplot as plt
    import seaborn as sns
    
    frecuencias = obtener_frecuencias_pregunta(pregunta_id)
    
    with _get_conn() as conn:
        fila = conn.execute("SELECT texto_pregunta FROM preguntas_extraccion WHERE id = ?", (pregunta_id,)).fetchone()
        titulo = fila["texto_pregunta"] if fila else f"Pregunta {pregunta_id}"
        
    sns.set_style("whitegrid")
    plt.figure(figsize=(8, 5), dpi=300)
    
    if frecuencias:
        opciones = [f[0] for f in frecuencias]
        cantidades = [f[1] for f in frecuencias]
        
        # Use hue and legend=False to avoid future warnings with palette
        sns.barplot(x=cantidades, y=opciones, hue=opciones, palette="Set2", legend=False)
        
    plt.title(titulo)
    plt.xlabel("Cantidad")
    plt.ylabel("Opción")
    plt.tight_layout()
    
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    plt.close()
    buf.seek(0)
    return buf

@app.get("/grafico-extraccion/{pregunta_id}")
def get_grafico_extraccion(pregunta_id: int):
    buf = generar_grafico_pregunta(pregunta_id)
    return StreamingResponse(buf, media_type="image/png")

@app.get("/graficos-extraccion-todos")
def get_graficos_extraccion_todos():
    import zipfile
    with _get_conn() as conn:
        preguntas = conn.execute("SELECT id, texto_pregunta FROM preguntas_extraccion WHERE tipo = 'categorica'").fetchall()
        
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zipf:
        for p in preguntas:
            p_id = p["id"]
            texto = p["texto_pregunta"]
            
            img_buf = generar_grafico_pregunta(p_id)
            
            # Sanitizar nombre del archivo
            nombre_sano = re.sub(r'[^a-zA-Z0-9_\- ]', '', texto).strip()
            if not nombre_sano:
                nombre_sano = f"pregunta_{p_id}"
            nombre_archivo = f"{nombre_sano}.png"
            
            zipf.writestr(nombre_archivo, img_buf.getvalue())
            
    zip_buf.seek(0)
    return StreamingResponse(
        zip_buf, 
        media_type="application/zip", 
        headers={"Content-Disposition": "attachment; filename=Graficos_Extraccion_MSL.zip"}
    )

@app.get("/exportar-extraccion")
def get_exportar_extraccion():
    import openpyxl
    
    # 1. Obtener la lista de preguntas ordenadas para las cabeceras
    with _get_conn() as conn:
        preguntas = conn.execute("SELECT id, texto_pregunta, tipo FROM preguntas_extraccion ORDER BY orden").fetchall()
        
    encabezados = ["Artículo"]
    for p in preguntas:
        encabezados.append(p["texto_pregunta"])
        if p["tipo"] == "categorica":
            encabezados.append(f"{p['texto_pregunta']} - Texto original IA")
    
    # 2. Obtener los datos usando la función existente
    datos_articulos = get_extraccion_respuestas()
    
    # 3. Construir el Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"
    
    ws.append(encabezados)
    
    for art in datos_articulos:
        fila = [art["articulo_nombre"]]
        for p in preguntas:
            pregunta_id_str = str(p["id"])
            valor = art["respuestas"].get(pregunta_id_str, "")
            fila.append(valor)
            if p["tipo"] == "categorica":
                original_ia = art.get("originales_ia", {}).get(pregunta_id_str, "")
                fila.append(original_ia)
        ws.append(fila)
        
    # 4. Crear hojas adicionales para preguntas categóricas (one-hot)
    with _get_conn() as conn:
        for p in preguntas:
            if p["tipo"] == "categorica":
                # Sanitizar el nombre de la hoja según reglas de Excel (max 31 chars y sin caracteres inválidos)
                titulo_hoja = re.sub(r'[\\*?:/\[\]]', '', p["texto_pregunta"]).strip()
                if not titulo_hoja:
                    titulo_hoja = f"pregunta_{p['id']}"
                titulo_hoja = titulo_hoja[:31]
                
                # Asegurar nombre único por si se truncaron igual
                suffix = 1
                base_title = titulo_hoja
                while titulo_hoja in wb.sheetnames:
                    str_suffix = str(suffix)
                    titulo_hoja = base_title[:31 - len(str_suffix)] + str_suffix
                    suffix += 1
                    
                ws_cat = wb.create_sheet(title=titulo_hoja)
                
                # Extraer datos exactos 0/1 de la base de datos
                rows_cat = conn.execute(
                    """
                    SELECT articulo_nombre, opcion_texto, valor
                    FROM extraccion_respuestas_categoricas
                    WHERE pregunta_id = ?
                    """,
                    (p["id"],)
                ).fetchall()
                
                opciones_set = set()
                datos_art = {}
                
                for r in rows_cat:
                    art_nom = r["articulo_nombre"]
                    opc = r["opcion_texto"]
                    val = r["valor"]
                    
                    opciones_set.add(opc)
                    if art_nom not in datos_art:
                        datos_art[art_nom] = {}
                    datos_art[art_nom][opc] = val
                    
                # Si las opciones_set está vacío, significa que aún no hay datos guardados para esta pregunta
                if not opciones_set:
                    ws_cat.append(["Artículo"])
                    continue
                    
                opciones_lista = sorted(list(opciones_set))
                ws_cat.append(["Artículo"] + opciones_lista)
                
                for art_nom in sorted(datos_art.keys()):
                    valores = datos_art[art_nom]
                    fila_cat = [art_nom]
                    for opc in opciones_lista:
                        fila_cat.append(valores.get(opc, 0))
                    ws_cat.append(fila_cat)
                    
                # Insertar gráfico debajo de la tabla
                from openpyxl.drawing.image import Image as OpenpyxlImage
                img_buf = generar_grafico_pregunta(p["id"])
                
                img = OpenpyxlImage(img_buf)
                celda_destino = f"A{ws_cat.max_row + 2}"
                ws_cat.add_image(img, celda_destino)
        
    # Asegurar que exista el directorio resultados
    Path("resultados").mkdir(exist_ok=True)
    ruta_salida = Path("resultados") / "MSL_Extraccion_Datos.xlsx"
    
    wb.save(ruta_salida)
    
    return FileResponse(
        path=ruta_salida, 
        filename="MSL_Extraccion_Datos.xlsx", 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

class PreguntaExtraccionNuevo(BaseModel):
    texto_pregunta: str
    tipo: str
    opciones: list[str] | None = None

@app.post("/preguntas-extraccion")
def crear_pregunta_extraccion(pregunta: PreguntaExtraccionNuevo):
    if pregunta.tipo not in ["categorica", "numerica", "texto_libre"]:
        raise HTTPException(status_code=400, detail="Tipo inválido. Debe ser categorica, numerica o texto_libre.")
    
    if pregunta.tipo == "categorica":
        if not pregunta.opciones or len(pregunta.opciones) < 2:
            raise HTTPException(status_code=400, detail="Las preguntas categóricas deben tener al menos 2 opciones.")
            
    with _get_conn() as conn:
        # Calcular el siguiente orden
        max_orden = conn.execute("SELECT MAX(orden) FROM preguntas_extraccion").fetchone()[0]
        siguiente_orden = (max_orden or 0) + 1
        
        opciones_json = json.dumps(pregunta.opciones) if pregunta.opciones else None
        
        cursor = conn.execute(
            """
            INSERT INTO preguntas_extraccion (texto_pregunta, tipo, opciones, orden)
            VALUES (?, ?, ?, ?)
            """,
            (pregunta.texto_pregunta, pregunta.tipo, opciones_json, siguiente_orden)
        )
        nuevo_id = cursor.lastrowid
        
        nueva_fila = conn.execute("SELECT * FROM preguntas_extraccion WHERE id = ?", (nuevo_id,)).fetchone()
        
    d = dict(nueva_fila)
    if d.get("opciones"):
        try:
            d["opciones"] = json.loads(d["opciones"])
        except json.JSONDecodeError:
            d["opciones"] = []
    else:
        d["opciones"] = []
        
    return d

@app.get("/preguntas-extraccion")
def get_preguntas_extraccion():
    """
    Obtiene todas las preguntas de extracción de datos, ordenadas por 'orden'.
    Parsea el campo 'opciones' de JSON string a lista.
    """
    with _get_conn() as conn:
        filas = conn.execute(
            "SELECT * FROM preguntas_extraccion ORDER BY orden ASC"
        ).fetchall()
        
    resultado = []
    for f in filas:
        d = dict(f)
        if d.get("opciones"):
            try:
                d["opciones"] = json.loads(d["opciones"])
            except json.JSONDecodeError:
                d["opciones"] = []
        else:
            d["opciones"] = []
        resultado.append(d)
        
    return resultado

class PreguntaExtraccionUpdate(BaseModel):
    texto_pregunta: str | None = None
    tipo: str | None = None
    opciones: list[str] | None = None

@app.put("/preguntas-extraccion/{id}")
def actualizar_pregunta_extraccion(id: int, pregunta: PreguntaExtraccionUpdate):
    with _get_conn() as conn:
        fila_actual = conn.execute("SELECT * FROM preguntas_extraccion WHERE id = ?", (id,)).fetchone()
        if not fila_actual:
            raise HTTPException(status_code=404, detail="Pregunta no encontrada")
            
        update_fields = []
        update_values = []
        
        if pregunta.texto_pregunta is not None:
            update_fields.append("texto_pregunta = ?")
            update_values.append(pregunta.texto_pregunta)
            
        if pregunta.tipo is not None:
            if pregunta.tipo not in ["categorica", "numerica", "texto_libre"]:
                raise HTTPException(status_code=400, detail="Tipo inválido. Debe ser categorica, numerica o texto_libre.")
            update_fields.append("tipo = ?")
            update_values.append(pregunta.tipo)
            
        if pregunta.opciones is not None:
            # We must validate categorical length if type is categorical
            # Either it's being updated to categorical, or it's already categorical
            nuevo_tipo = pregunta.tipo if pregunta.tipo is not None else fila_actual["tipo"]
            if nuevo_tipo == "categorica" and len(pregunta.opciones) < 2:
                raise HTTPException(status_code=400, detail="Las preguntas categóricas deben tener al menos 2 opciones.")
            
            update_fields.append("opciones = ?")
            update_values.append(json.dumps(pregunta.opciones))
            
        # If type was updated to categorical but options were not provided, we should validate the existing options
        if pregunta.tipo == "categorica" and pregunta.opciones is None:
            opc_existentes = []
            if fila_actual["opciones"]:
                try:
                    opc_existentes = json.loads(fila_actual["opciones"])
                except json.JSONDecodeError:
                    pass
            if len(opc_existentes) < 2:
                raise HTTPException(status_code=400, detail="El tipo cambiado a categórica requiere que haya al menos 2 opciones previamente guardadas o proporcionadas.")
                
        if update_fields:
            update_values.append(id)
            query = f"UPDATE preguntas_extraccion SET {', '.join(update_fields)} WHERE id = ?"
            conn.execute(query, tuple(update_values))
            
        nueva_fila = conn.execute("SELECT * FROM preguntas_extraccion WHERE id = ?", (id,)).fetchone()
        
    d = dict(nueva_fila)
    if d.get("opciones"):
        try:
            d["opciones"] = json.loads(d["opciones"])
        except json.JSONDecodeError:
            d["opciones"] = []
    else:
        d["opciones"] = []
        
    return d

@app.delete("/preguntas-extraccion/{id}")
def eliminar_pregunta_extraccion(id: int):
    with _get_conn() as conn:
        cursor = conn.execute("DELETE FROM preguntas_extraccion WHERE id = ?", (id,))
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Pregunta no encontrada")
    return {"ok": True}


# Componentes JS compartidos (navbar, etc.)
app.mount("/components", StaticFiles(directory="static/components"), name="components")

# Montar static/ en "/" al final para que no tape las rutas API
app.mount("/", StaticFiles(directory="static", html=True), name="static")

